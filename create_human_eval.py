#!/usr/bin/env python3
"""
ULAB — Human Evaluation Excel Generator
Создаёт Excel-файл для ручной оценки ответов AI-моделей экспертами.
Названия моделей скрыты (слепая оценка).

Использование:
    python3 create_human_eval.py
"""

import json
import random
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Установите зависимость: pip3 install openpyxl")
    exit(1)

BASE_DIR    = Path(__file__).parent
RESULTS_JSON = BASE_DIR / "results_raw.json"

# ── Цвета ──────────────────────────────────────────────────────
C = {
    "primary":   "1B3A6B",
    "accent":    "2E7CF6",
    "bg":        "F4F6FA",
    "white":     "FFFFFF",
    "border":    "CBD5E1",
    "even":      "F8FAFC",
    "odd":       "FFFFFF",
    "score_bg":  "EFF6FF",
    "score_fg":  "1D4ED8",
    "sep":       "E2E8F0",
    "warn":      "FFF3F3",
    "warn_txt":  "CC0000",
    # Регистры
    "formal_hdr":   "1D4ED8",
    "formal_bg":    "DBEAFE",
    "informal_hdr": "15803D",
    "informal_bg":  "DCFCE7",
    "slang_hdr":    "B45309",
    "slang_bg":     "FEF9C3",
    # Шкала
    "s5": "1A9850",
    "s4": "91CF60",
    "s3": "FEE08B",
    "s2": "FC8D59",
    "s1": "D73027",
}

REG = {
    "formal_business": {
        "label": "Официальный деловой стиль",
        "sheet": "🏦 Официальный стиль",
        "hdr":   C["formal_hdr"],
        "bg":    C["formal_bg"],
    },
    "informal": {
        "label": "Повседневный стиль",
        "sheet": "💬 Повседневный",
        "hdr":   C["informal_hdr"],
        "bg":    C["informal_bg"],
    },
    "slang": {
        "label": "Разговорный стиль (слэнг)",
        "sheet": "🗣 Разговорный",
        "hdr":   C["slang_hdr"],
        "bg":    C["slang_bg"],
    },
}


# ── Стили ──────────────────────────────────────────────────────
def border(color=None):
    s = Side(style="thin", color=color or C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def fill(color):
    return PatternFill("solid", fgColor=color)

def font(bold=False, size=10, color=None, italic=False):
    return Font(name="Calibri", bold=bold, size=size,
                color=color or "1A1A2E", italic=italic)

def align(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


def cell_set(c, value="", fg=None, ft=None, al=None, bd=None):
    c.value    = value
    if fg: c.fill      = fg
    if ft: c.font      = ft
    if al: c.alignment = al
    if bd: c.border    = bd


# ═══════════════════════════════════════════════════════════════
# ЛИСТ: ИНСТРУКЦИЯ
# ═══════════════════════════════════════════════════════════════
def build_instruction(ws):
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 16

    r = 1

    # Заголовок
    ws.merge_cells(f"A{r}:D{r}")
    cell_set(ws.cell(r, 1),
             "ULAB — Экспертная оценка ответов AI-моделей на узбекском языке",
             fill(C["primary"]),
             font(bold=True, size=15, color=C["white"]),
             align("center", "center"))
    ws.row_dimensions[r].height = 44
    r += 1

    ws.merge_cells(f"A{r}:D{r}")
    cell_set(ws.cell(r, 1), "Инструкция для эксперта",
             fill(C["bg"]),
             font(bold=True, size=12, color=C["accent"]),
             align("center", "center"))
    ws.row_dimensions[r].height = 28
    r += 2

    # ── Задание ─────────────────────────────────────────────────
    ws.merge_cells(f"B{r}:D{r}")
    ws.cell(r, 2).value = "ЧТО НУЖНО СДЕЛАТЬ"
    ws.cell(r, 2).font  = font(bold=True, size=11, color=C["primary"])
    r += 1

    tasks = [
        "1.  Перейдите на нужный лист: «Официальный стиль», «Повседневный» или «Разговорный».",
        "2.  Прочитайте вопрос — он выделен тёмно-синей строкой.",
        "3.  Прочитайте ответ каждой модели (М-01, М-02 и т.д.).",
        "4.  Поставьте оценку от 1 до 5 в ячейки D1, D2, D3, D4 для каждой модели.",
        "5.  При желании оставьте краткий комментарий в последнем столбце.",
        "6.  Названия моделей скрыты специально — оценивайте только качество текста, не угадывайте.",
    ]
    for t in tasks:
        ws.merge_cells(f"B{r}:D{r}")
        cell_set(ws.cell(r, 2), t,
                 fill(C["even"]) if r % 2 == 0 else fill(C["odd"]),
                 font(size=10),
                 align("left", "center", wrap=True, indent=1),
                 border())
        ws.row_dimensions[r].height = 20
        r += 1
    r += 1

    # ── Критерии ────────────────────────────────────────────────
    ws.merge_cells(f"B{r}:D{r}")
    ws.cell(r, 2).value = "КРИТЕРИИ ОЦЕНКИ"
    ws.cell(r, 2).font  = font(bold=True, size=11, color=C["primary"])
    r += 1

    # Шапка
    for col, hdr in [(2, "Критерий"), (3, "Что оценивается"), (4, "Вес")]:
        cell_set(ws.cell(r, col), hdr,
                 fill(C["primary"]),
                 font(bold=True, size=10, color=C["white"]),
                 align("center", "center"),
                 border())
    ws.row_dimensions[r].height = 22
    r += 1

    criteria = [
        ("D1 — Точность ответа",    "Ответ правильно и полно отвечает на заданный вопрос",                         "25%"),
        ("D2 — Качество языка",     "Грамматика, орфография, правильность узбекского языка",                       "25%"),
        ("D3 — Соответствие стилю", "Стиль ответа соответствует регистру (деловой / повседневный / разговорный)",  "25%"),
        ("D4 — Естественность",     "Ответ звучит как живой человек, а не как машинный перевод",                   "15%"),
    ]
    for i, (crit, desc, w) in enumerate(criteria):
        bg = fill(C["even"]) if i % 2 == 0 else fill(C["odd"])
        cell_set(ws.cell(r, 2), crit, bg, font(bold=True, size=10), align("left", "center", indent=1), border())
        cell_set(ws.cell(r, 3), desc, bg, font(size=10),            align("left", "center", wrap=True, indent=1), border())
        cell_set(ws.cell(r, 4), w,    bg, font(bold=True, size=10), align("center", "center"), border())
        ws.row_dimensions[r].height = 28
        r += 1
    r += 1

    # ── Шкала ───────────────────────────────────────────────────
    ws.merge_cells(f"B{r}:D{r}")
    ws.cell(r, 2).value = "ШКАЛА ОЦЕНОК"
    ws.cell(r, 2).font  = font(bold=True, size=11, color=C["primary"])
    r += 1

    for col, hdr in [(2, "Балл"), (3, "Значение"), (4, "Пример")]:
        cell_set(ws.cell(r, col), hdr,
                 fill(C["primary"]),
                 font(bold=True, size=10, color=C["white"]),
                 align("center", "center"),
                 border())
    ws.row_dimensions[r].height = 22
    r += 1

    scale = [
        ("5", "Отлично",     "Точный, грамотный, полностью соответствует стилю",  C["s5"]),
        ("4", "Хорошо",      "Небольшие недочёты, но в целом качественно",         C["s4"]),
        ("3", "Средне",      "Есть заметные ошибки или неточности",                C["s3"]),
        ("2", "Плохо",       "Много ошибок, ответ слабый",                         C["s2"]),
        ("1", "Очень плохо", "Не по теме, непонятный или неправильный",            C["s1"]),
    ]
    for i, (score, label, example, sc) in enumerate(scale):
        bg = fill(C["even"]) if i % 2 == 0 else fill(C["odd"])
        cell_set(ws.cell(r, 2), score,   fill(sc), font(bold=True, size=12, color=C["white"]), align("center", "center"), border())
        cell_set(ws.cell(r, 3), label,   bg,        font(bold=True, size=10),                   align("left", "center", indent=1), border())
        cell_set(ws.cell(r, 4), example, bg,        font(size=9, color="555555"),               align("left", "center", wrap=True, indent=1), border())
        ws.row_dimensions[r].height = 24
        r += 1
    r += 1

    # Предупреждение
    ws.merge_cells(f"B{r}:D{r}")
    cell_set(ws.cell(r, 2),
             "⚠  Важно: заполните все 4 оценки (D1–D4) для каждой модели. "
             "Итоговый балл рассчитывается автоматически.",
             fill(C["warn"]),
             font(bold=True, size=10, color=C["warn_txt"]),
             align("left", "center", wrap=True, indent=1))
    ws.row_dimensions[r].height = 30


# ═══════════════════════════════════════════════════════════════
# ЛИСТ: ОЦЕНКА (один на регистр)
# ═══════════════════════════════════════════════════════════════
def build_eval_sheet(ws, questions, model_codes, register, seed):
    rng  = REG[register]
    hdr_color = rng["hdr"]
    bg_color  = rng["bg"]

    # Ширины
    widths = {1: 5, 2: 10, 3: 58, 4: 7, 5: 7, 6: 7, 7: 7, 8: 11, 9: 36}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    r = 1

    # Заголовок листа
    ws.merge_cells(f"A{r}:I{r}")
    cell_set(ws.cell(r, 1),
             f"ULAB — {rng['label'].upper()}   |   Слепая экспертная оценка",
             fill(hdr_color),
             font(bold=True, size=13, color=C["white"]),
             align("center", "center"))
    ws.row_dimensions[r].height = 34
    r += 1

    # Шапка таблицы
    headers = ["№", "Модель", "Текст вопроса / Ответ модели",
               "D1\n(1–5)", "D2\n(1–5)", "D3\n(1–5)", "D4\n(1–5)",
               "Итого\n(авто)", "Комментарий эксперта"]
    header_row = r
    for col, hdr in enumerate(headers, 1):
        cell_set(ws.cell(r, col), hdr,
                 fill(C["primary"]),
                 font(bold=True, size=10, color=C["white"]),
                 align("center", "center", wrap=True),
                 border())
    ws.row_dimensions[r].height = 36
    r += 1

    # Валидация: только 1-5 в столбцах D1-D4 (cols 4-7)
    dv = DataValidation(
        type="whole", operator="between", formula1="1", formula2="5",
        error="Введите целое число от 1 до 5",
        errorTitle="Неверное значение",
        prompt="Оценка от 1 до 5",
        promptTitle="Введите оценку",
        showErrorMessage=True,
    )
    dv.sqref = f"D{r}:G{r + len(questions) * 20}"
    ws.add_data_validation(dv)

    n = 1
    rnd = random.Random(seed)  # фиксированный seed для воспроизводимости

    for q_idx, q in enumerate(questions):
        # Перемешать модели для каждого вопроса (разный порядок)
        model_ids_shuffled = rnd.sample(list(model_codes.keys()), len(model_codes))

        # ── Строка вопроса ─────────────────────────────────────
        ws.merge_cells(f"A{r}:B{r}")
        ws.merge_cells(f"C{r}:I{r}")

        cell_set(ws.cell(r, 1), f"{q['id']}",
                 fill(C["primary"]),
                 font(bold=True, size=10, color=C["white"]),
                 align("center", "center"),
                 border())
        cell_set(ws.cell(r, 3), f"❓  {q['text']}",
                 fill(C["primary"]),
                 font(bold=True, size=10, color=C["white"]),
                 align("left", "center", wrap=True, indent=1),
                 border())
        ws.row_dimensions[r].height = max(36, min(len(q["text"]) // 2, 80))
        r += 1

        # ── Строки моделей ─────────────────────────────────────
        for m_idx, model_id in enumerate(model_ids_shuffled):
            code     = model_codes[model_id]
            resp_obj = q.get("responses", {}).get(model_id, {})
            resp_txt = resp_obj.get("response", "")

            row_bg = fill(C["even"]) if m_idx % 2 == 0 else fill(C["odd"])

            # № строки
            cell_set(ws.cell(r, 1), n,
                     row_bg, font(size=9, color="888888"),
                     align("center", "top"), border())

            # Код модели
            cell_set(ws.cell(r, 2), code,
                     fill(bg_color),
                     font(bold=True, size=10, color=hdr_color),
                     align("center", "top"), border())

            # Текст ответа
            if resp_txt:
                cell_set(ws.cell(r, 3), resp_txt,
                         row_bg, font(size=9),
                         align("left", "top", wrap=True, indent=1), border())
            else:
                cell_set(ws.cell(r, 3), "(ответ отсутствует)",
                         row_bg, font(size=9, italic=True, color="AAAAAA"),
                         align("left", "top", indent=1), border())

            # D1–D4 (пустые, для заполнения)
            for col in [4, 5, 6, 7]:
                cell_set(ws.cell(r, col), "",
                         fill(C["score_bg"]),
                         font(bold=True, size=11, color=C["score_fg"]),
                         align("center", "center"), border())

            # Итого (формула)
            d1 = get_column_letter(4) + str(r)
            d2 = get_column_letter(5) + str(r)
            d3 = get_column_letter(6) + str(r)
            d4 = get_column_letter(7) + str(r)
            formula = f'=IFERROR(IF(AND({d1}<>"",{d2}<>"",{d3}<>"",{d4}<>""),ROUND(({d1}*25+{d2}*25+{d3}*25+{d4}*15-90)/360*100,0),""  ),"")'
            cell_set(ws.cell(r, 8), formula,
                     fill("E0ECFF"),
                     font(bold=True, size=10, color=C["score_fg"]),
                     align("center", "center"), border())

            # Комментарий
            cell_set(ws.cell(r, 9), "",
                     row_bg, font(size=9, italic=True, color="888888"),
                     align("left", "top", wrap=True, indent=1), border())

            # Высота строки зависит от длины ответа
            ws.row_dimensions[r].height = max(50, min(len(resp_txt) // 5, 180))
            r += 1
            n += 1

        # Разделитель между вопросами
        for col in range(1, 10):
            ws.cell(r, col).fill = fill(C["sep"])
        ws.row_dimensions[r].height = 5
        r += 1

    # Закрепить заголовок
    ws.freeze_panes = f"A{header_row + 1}"


# ═══════════════════════════════════════════════════════════════
# ЛИСТ: КЛЮЧ (скрытый служебный)
# ═══════════════════════════════════════════════════════════════
def build_key_sheet(ws, model_ids, model_codes, models_meta):
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16

    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    cell_set(ws.cell(r, 1), "🔑 Ключ — Расшифровка кодов моделей",
             fill(C["primary"]),
             font(bold=True, size=14, color=C["white"]),
             align("center", "center"))
    ws.row_dimensions[r].height = 36
    r += 1

    ws.merge_cells(f"A{r}:D{r}")
    cell_set(ws.cell(r, 1),
             "Этот лист только для организатора. Эксперты не должны смотреть сюда до окончания оценки.",
             fill("FFF8E1"),
             font(bold=False, size=10, color="7B5E00"),
             align("center", "center", wrap=True))
    ws.row_dimensions[r].height = 28
    r += 2

    # Шапка
    for col, hdr in enumerate(["Код", "Модель", "Провайдер", "Тип"], 1):
        cell_set(ws.cell(r, col), hdr,
                 fill(C["primary"]),
                 font(bold=True, size=10, color=C["white"]),
                 align("center", "center"), border())
    ws.row_dimensions[r].height = 22
    r += 1

    for i, mid in enumerate(model_ids):
        code = model_codes[mid]
        meta = models_meta.get(mid, {})
        bg = fill(C["even"]) if i % 2 == 0 else fill(C["odd"])
        vals = [
            (code,                               True,  "center"),
            (meta.get("name", mid),              True,  "left"),
            (meta.get("provider", "—"),          False, "left"),
            ("Коммерческая" if meta.get("type") == "commercial" else "Open-Source", False, "center"),
        ]
        for col, (v, bold, h) in enumerate(vals, 1):
            cell_set(ws.cell(r, col), v, bg,
                     font(bold=bold, size=10),
                     align(h, "center", indent=1 if h == "left" else 0),
                     border())
        ws.row_dimensions[r].height = 20
        r += 1


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════
def main():
    if not RESULTS_JSON.exists():
        print("  ОШИБКА: results_raw.json не найден.")
        print("  Сначала запустите: python3 run_benchmark.py && python3 score_responses.py")
        return

    data      = json.loads(RESULTS_JSON.read_text(encoding="utf-8"))
    questions = data["questions"]
    models    = data["models"]

    # Анонимизация: перемешиваем модели один раз и назначаем коды
    model_ids = list(models.keys())
    random.shuffle(model_ids)
    model_codes = {mid: f"М-{i + 1:02d}" for i, mid in enumerate(model_ids)}

    # Seed для воспроизводимости порядка внутри листов (фиксируем дату)
    seed = int(datetime.now().strftime("%Y%m%d"))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Лист 1 — Инструкция
    ws_instr = wb.create_sheet("📋 Инструкция")
    build_instruction(ws_instr)

    # Листы 2–4 — Оценка по регистрам
    for reg in ["formal_business", "informal", "slang"]:
        reg_qs = [q for q in questions if q["register"] == reg]
        ws     = wb.create_sheet(REG[reg]["sheet"])
        build_eval_sheet(ws, reg_qs, model_codes, reg, seed)

    # Лист 5 — Ключ (видимый, в конце книги)
    ws_key = wb.create_sheet("🔑 Ключ моделей")
    build_key_sheet(ws_key, model_ids, model_codes, models)

    # Сохранить
    date_str = datetime.now().strftime("%Y-%m-%d")
    out_file = BASE_DIR / f"ULAB_human_eval_{date_str}.xlsx"
    wb.save(out_file)

    total_cells = len(model_ids) * len(questions)
    print(f"\n{'═'*60}")
    print(f"  ULAB — Excel для экспертной оценки создан")
    print(f"{'═'*60}")
    print(f"\n  Файл:       {out_file.name}")
    print(f"  Моделей:    {len(model_ids)}  (коды М-01 — М-{len(model_ids):02d})")
    print(f"  Вопросов:   {len(questions)}")
    print(f"  Ответов:    {total_cells}  (по 3 листам)")
    print(f"\n  Листы:")
    print(f"    📋 Инструкция        — прочитать перед началом")
    print(f"    🏦 Официальный стиль — 20 вопросов × {len(model_ids)} моделей")
    print(f"    💬 Повседневный      — 20 вопросов × {len(model_ids)} моделей")
    print(f"    🗣 Разговорный       — 20 вопросов × {len(model_ids)} моделей")
    print(f"    🔑 Ключ моделей      — последний лист (для организатора)")
    print(f"\n{'─'*60}")
    print(f"  КЛЮЧ — расшифровка кодов (сохраните!)")
    print(f"{'─'*60}")
    for mid in model_ids:
        code = model_codes[mid]
        meta = models.get(mid, {})
        name = meta.get("name", mid)
        prov = meta.get("provider", "")
        kind = "Коммерческая" if meta.get("type") == "commercial" else "Open-Source"
        print(f"  {code}  →  {name:<22} ({prov}, {kind})")
    print(f"{'─'*60}")
    print(f"\n  Порядок моделей перемешан на каждый вопрос.")
    print(f"  Итоговый балл считается автоматически по формуле.\n")


if __name__ == "__main__":
    main()
